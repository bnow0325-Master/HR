# -*- coding: utf-8 -*-
from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(r"C:\Users\Elon Choo PC\Documents\출퇴근관리 프로그램")
LEAVE_XLSX = Path(r"C:\Users\Elon Choo PC\Downloads\annualLeaveList.xlsx")
COMMUTE_XLSX = Path(r"C:\Users\Elon Choo PC\Downloads\commuteList.xlsx")
OUT_SQL = ROOT / "sql" / "checkinout_seed.sql"

CUTOFF = datetime(2026, 7, 21)
LEE_RESIGNED = datetime(2026, 7, 20)
BASE_YEAR = 2026


def sql_nvarchar(value: str | None) -> str:
    if value is None:
        return "NULL"
    return "N'" + value.replace("'", "''") + "'"


def sql_varchar(value: str | None) -> str:
    if value is None:
        return "NULL"
    return "'" + value.replace("'", "''") + "'"


def sql_date(value: str | None) -> str:
    if not value:
        return "NULL"
    normalized = datetime.strptime(value, "%Y%m%d").strftime("%Y-%m-%d")
    return f"CAST('{normalized}' AS date)"


def sql_datetime(value: datetime) -> str:
    return f"CAST('{value.strftime('%Y-%m-%d %H:%M:%S')}' AS datetime2(0))"


def parse_method(raw: str) -> str | None:
    if "(" not in raw or ")" not in raw:
        return None
    return raw.split("(", 1)[1].split(")", 1)[0].strip() or None


def parse_timestamp(date_raw: Any, time_raw: str) -> datetime:
    day = datetime.strptime(str(date_raw)[:8], "%Y%m%d")
    hhmm = time_raw.split("(", 1)[0]
    hour, minute = hhmm.split(":")
    return day.replace(hour=int(hour), minute=int(minute), second=0, microsecond=0)


def load_leave_rows() -> list[dict[str, Any]]:
    wb = load_workbook(LEAVE_XLSX, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[1]:
            continue
        rows.append(
            {
                "leave_group": row[0],
                "name": row[1],
                "login_id": row[2],
                "department": row[3],
                "hire_date": str(row[4]) if row[4] else None,
                "granted_days": float(row[5] or 0),
                "under_one_granted_days": float(row[6] or 0),
                "under_one_carryover_days": float(row[7] or 0),
                "carryover_days": float(row[8] or 0),
                "used_days": float(row[9] or 0),
                "adjusted_days": float(row[10] or 0),
                "remaining_days": float(row[11] or 0),
                "is_active": row[1] != "이세언",
                "resigned_on": "20260720" if row[1] == "이세언" else None,
            }
        )
    return rows


def load_commute_people() -> dict[str, dict[str, Any]]:
    wb = load_workbook(COMMUTE_XLSX, data_only=True)
    ws = wb[wb.sheetnames[0]]
    people: dict[str, dict[str, Any]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[1]
        if not name or name in people:
            continue
        people[name] = {
            "name": name,
            "login_id": row[2],
            "department": row[3],
        }
    return people


def load_attendance_rows() -> list[dict[str, Any]]:
    wb = load_workbook(COMMUTE_XLSX, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[1]
        date_raw = row[4]
        if not name or not date_raw:
            continue
        day = datetime.strptime(str(date_raw)[:8], "%Y%m%d")
        if day > CUTOFF:
            continue
        if name == "이세언" and day > LEE_RESIGNED:
            continue

        check_in = row[7]
        check_out = row[8]

        if check_in:
            rows.append(
                {
                    "name": name,
                    "record_type": "IN",
                    "recorded_at": parse_timestamp(date_raw, check_in),
                    "record_method": parse_method(check_in),
                }
            )
        if check_out:
            rows.append(
                {
                    "name": name,
                    "record_type": "OUT",
                    "recorded_at": parse_timestamp(date_raw, check_out),
                    "record_method": parse_method(check_out),
                }
            )
    return rows


def build_employee_rows() -> list[dict[str, Any]]:
    leave_rows = {row["name"]: row for row in load_leave_rows()}
    commute_people = load_commute_people()
    names = sorted(set(leave_rows) | set(commute_people))
    employees: list[dict[str, Any]] = []
    for name in names:
        leave = leave_rows.get(name, {})
        commute = commute_people.get(name, {})
        login_id = leave.get("login_id") or commute.get("login_id")
        employee_code = login_id or name
        employees.append(
            {
                "employee_code": employee_code,
                "employee_name": name,
                "login_id": login_id,
                "department_name": leave.get("department") or commute.get("department"),
                "leave_group": leave.get("leave_group"),
                "hire_date": leave.get("hire_date"),
                "is_active": leave.get("is_active", True),
                "resigned_on": leave.get("resigned_on"),
            }
        )
    return employees


def render_sql() -> str:
    employees = build_employee_rows()
    leave_rows = {row["name"]: row for row in load_leave_rows()}
    attendance_rows = load_attendance_rows()

    employee_names = [row["employee_name"] for row in employees]
    delete_codes = ", ".join(sql_nvarchar(row["employee_code"]) for row in employees)

    lines: list[str] = [
        "SET NOCOUNT ON;",
        "",
        "IF COL_LENGTH('dbo.CHECKINOUT_EMPLOYEE', 'LOGIN_ID') IS NULL",
        "    ALTER TABLE dbo.CHECKINOUT_EMPLOYEE ADD LOGIN_ID NVARCHAR(255) NULL;",
        "IF COL_LENGTH('dbo.CHECKINOUT_EMPLOYEE', 'HIRED_ON') IS NULL",
        "    ALTER TABLE dbo.CHECKINOUT_EMPLOYEE ADD HIRED_ON DATE NULL;",
        "IF COL_LENGTH('dbo.CHECKINOUT_EMPLOYEE', 'LEAVE_GROUP') IS NULL",
        "    ALTER TABLE dbo.CHECKINOUT_EMPLOYEE ADD LEAVE_GROUP NVARCHAR(100) NULL;",
        "IF COL_LENGTH('dbo.CHECKINOUT_EMPLOYEE', 'RESIGNED_ON') IS NULL",
        "    ALTER TABLE dbo.CHECKINOUT_EMPLOYEE ADD RESIGNED_ON DATE NULL;",
        "",
    ]

    for employee in employees:
        lines.extend(
            [
                "MERGE dbo.CHECKINOUT_EMPLOYEE AS target",
                "USING (",
                "    SELECT",
                f"        {sql_nvarchar(employee['employee_code'])} AS EMPLOYEE_CODE,",
                f"        {sql_nvarchar(employee['employee_name'])} AS EMPLOYEE_NAME,",
                f"        {sql_nvarchar(employee['department_name'])} AS DEPARTMENT_NAME,",
                f"        {sql_nvarchar(employee['login_id'])} AS LOGIN_ID,",
                f"        {sql_nvarchar(employee['leave_group'])} AS LEAVE_GROUP,",
                f"        {sql_date(employee['hire_date'])} AS HIRED_ON,",
                f"        {1 if employee['is_active'] else 0} AS IS_ACTIVE,",
                f"        {sql_date(employee['resigned_on'])} AS RESIGNED_ON",
                ") AS source",
                "ON target.EMPLOYEE_CODE = source.EMPLOYEE_CODE",
                "WHEN MATCHED THEN UPDATE SET",
                "    EMPLOYEE_NAME = source.EMPLOYEE_NAME,",
                "    DEPARTMENT_NAME = source.DEPARTMENT_NAME,",
                "    LOGIN_ID = source.LOGIN_ID,",
                "    LEAVE_GROUP = source.LEAVE_GROUP,",
                "    HIRED_ON = source.HIRED_ON,",
                "    IS_ACTIVE = source.IS_ACTIVE,",
                "    RESIGNED_ON = source.RESIGNED_ON,",
                "    UPDATED_AT = SYSDATETIME()",
                "WHEN NOT MATCHED THEN INSERT (",
                "    EMPLOYEE_CODE, EMPLOYEE_NAME, DEPARTMENT_NAME, LOGIN_ID, LEAVE_GROUP, HIRED_ON, IS_ACTIVE, RESIGNED_ON",
                ") VALUES (",
                "    source.EMPLOYEE_CODE, source.EMPLOYEE_NAME, source.DEPARTMENT_NAME, source.LOGIN_ID, source.LEAVE_GROUP, source.HIRED_ON, source.IS_ACTIVE, source.RESIGNED_ON",
                ");",
                "",
            ]
        )

    lines.extend(
        [
            f"DELETE ar FROM dbo.CHECKINOUT_ATTENDANCE_RECORD ar INNER JOIN dbo.CHECKINOUT_EMPLOYEE e ON e.CHECKINOUT_EMPLOYEE_ID = ar.CHECKINOUT_EMPLOYEE_ID WHERE e.EMPLOYEE_CODE IN ({delete_codes}) AND ar.RECORDED_AT < CAST('2026-07-22 00:00:00' AS datetime2(0));",
            "",
        ]
    )

    for row in attendance_rows:
        employee_code = next(
            employee["employee_code"]
            for employee in employees
            if employee["employee_name"] == row["name"]
        )
        lines.extend(
            [
                "INSERT INTO dbo.CHECKINOUT_ATTENDANCE_RECORD (",
                "    CHECKINOUT_EMPLOYEE_ID, RECORD_TYPE, RECORDED_AT, RECORD_METHOD, IS_VERIFIED, CREATED_AT",
                ")",
                "SELECT",
                "    e.CHECKINOUT_EMPLOYEE_ID,",
                f"    {sql_varchar(row['record_type'])},",
                f"    {sql_datetime(row['recorded_at'])},",
                f"    {sql_nvarchar(row['record_method'])},",
                "    0,",
                f"    {sql_datetime(row['recorded_at'])}",
                "FROM dbo.CHECKINOUT_EMPLOYEE e",
                f"WHERE e.EMPLOYEE_CODE = {sql_nvarchar(employee_code)};",
                "",
            ]
        )

    for employee in employees:
        leave = leave_rows.get(employee["employee_name"])
        if not leave:
            continue
        granted_total = (
            leave["granted_days"]
            + leave["under_one_granted_days"]
            + leave["under_one_carryover_days"]
            + leave["carryover_days"]
        )
        lines.extend(
            [
                "MERGE dbo.CHECKINOUT_ANNUAL_LEAVE_BALANCE AS target",
                "USING (",
                "    SELECT",
                f"        {sql_nvarchar(employee['employee_code'])} AS EMPLOYEE_CODE,",
                f"        {BASE_YEAR} AS BASE_YEAR,",
                f"        CAST({granted_total:.2f} AS decimal(5,2)) AS GRANTED_DAYS,",
                f"        CAST({leave['used_days']:.2f} AS decimal(5,2)) AS USED_DAYS,",
                f"        CAST({leave['adjusted_days']:.2f} AS decimal(5,2)) AS ADJUSTED_DAYS",
                ") AS source",
                "ON target.CHECKINOUT_EMPLOYEE_ID = (SELECT CHECKINOUT_EMPLOYEE_ID FROM dbo.CHECKINOUT_EMPLOYEE WHERE EMPLOYEE_CODE = source.EMPLOYEE_CODE)",
                "AND target.BASE_YEAR = source.BASE_YEAR",
                "WHEN MATCHED THEN UPDATE SET",
                "    GRANTED_DAYS = source.GRANTED_DAYS,",
                "    USED_DAYS = source.USED_DAYS,",
                "    ADJUSTED_DAYS = source.ADJUSTED_DAYS,",
                "    UPDATED_AT = SYSDATETIME()",
                "WHEN NOT MATCHED THEN INSERT (",
                "    CHECKINOUT_EMPLOYEE_ID, BASE_YEAR, GRANTED_DAYS, USED_DAYS, ADJUSTED_DAYS",
                ") VALUES (",
                "    (SELECT CHECKINOUT_EMPLOYEE_ID FROM dbo.CHECKINOUT_EMPLOYEE WHERE EMPLOYEE_CODE = source.EMPLOYEE_CODE),",
                "    source.BASE_YEAR, source.GRANTED_DAYS, source.USED_DAYS, source.ADJUSTED_DAYS",
                ");",
                "",
            ]
        )

    return "\n".join(lines)


def main() -> None:
    OUT_SQL.parent.mkdir(parents=True, exist_ok=True)
    OUT_SQL.write_text(render_sql(), encoding="utf-8")
    print(f"Wrote {OUT_SQL}")


if __name__ == "__main__":
    main()
